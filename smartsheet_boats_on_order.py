#!/usr/bin/env python3
"""
convert smartsheet boats on order reports to specially formatted execl sheets
and pdfs

pyinstaller --onefile smartsheet.spec smartsheet_boats_on_order.py

To do:
    port from OpenPyXL to XlsxWriter for RichText capability
"""
import smartsheet
import datetime
import os
import shutil
import sys
import contextlib
import subprocess
import openpyxl
import dateparser
import datedelta
import click
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from dotenv import load_dotenv
from emailer import mail_results
from PyPDF2 import PdfFileReader, PdfFileWriter

api = ''
source_dir = ''
target_dir = ''
rollover = 0
one_date_fmt = ''
two_date_fmt = ''
log_text = ''
errors = False
is_pdf = False
date_coloring = True
temp_date = None


# =========================================================
# column class for formatting rules
# =========================================================
class Column:
    name = 'Arial'
    size = '9'
    bold = False
    italic = False
    color = '000000'
    bg_color = 'FFFFFF'

    def __init__(self, old, new, title, function):
        self.info = {}
        self.info['text'] = ''
        self.info['old'] = old
        self.info['new'] = new
        self.info['title'] = title
        self.reset()
        self.function = function

    def reset(self):
        self.info['text'] = ''
        self.info['name'] = ''
        self.info['size'] = ''
        self.info['bold'] = False
        self.info['italic'] = False
        self.info['color'] = ''
        self.info['bg_color'] = ''

    def font(self):
        return Font(
            name=self.info['name'] or Column.name,
            size=self.info['size'] or Column.size,
            bold=self.info['bold'] or Column.bold,
            italic=self.info['italic'] or Column.italic,
            color=self.info['color'] or Column.color
        )

    def bg(self):
        return self.info['bg_color'] or Column.bg_color

    def run(self):
        self.info = self.function(self.info)


# =========================================================
# column formatting logic
# =========================================================
def noop(info):
    """
    default do nothing column
    """
    return info


def boat_model(info):
    """
    Change background color on OS and HardTops Mods super() to affect all
    columns that do not have individual overrides
    """
    info['size'] = 8
    Column.bg_color = 'FFFFFF'
    if info['text'].find("OS") != -1:
        Column.bg_color = 'FFA6A6A6'
    if info['text'].replace(" ", "").lower().find('hardtop') != -1:
        Column.bg_color = 'FFD9D9D9'
    return info


def hull_space(info):
    """
    Add a space before the hull number
    """
    if info['text'] and is_pdf:
        info['text'] = ' ' + info['text']
    return info


def colors_interior(info):
    """
    Set font size to 8 on color interior/exterior column
    """
    info['size'] = 8
    return info


def order_details(info):
    """
    set background of colum to orange if there is no text
    """
    if not info['text']:
        info['bg_color'] = 'FFFFC000'
    return info


def start_finish(info):
    """
    format date as Jan/Feb or January and apply coloring
      red for current month
      blue for next month
      black for all other months
    """
    if temp_date:
        info['text'] = temp_date
    info['text'], Column.color = start_info(info['text'])
    return info


def current_phase(info):
    """
    if phase contains any of the following phrases then
    replace all text with that phrase otherwise no text at all
    """
    phases = [
        'Waiting Production',
        'Pre-Fab',
        'Fab',
        'Upholstery',
        'Paint',
        'Outfitting',
        'Trials',
        'Completed',
        'Delivered',
    ]
    text = info['text'].lower()
    info['text'] = ''
    for phase in phases:
        if text.find(phase.lower()) != -1:
            info['text'] = phase
            break
    return info

def set_start_date(info):
    global temp_date
    try:
        temp_date = datetime.datetime.strptime(info['text'],'%m/%d/%y').strftime("%b %d")
    except ValueError:
        temp_date = None
    return info

def set_end_date(info):
    global temp_date
    if temp_date == None:
        return info
    try:
        temp_date = temp_date + " / " + datetime.datetime.strptime(info['text'],'%m/%d/%y').strftime("%b %d")
    except ValueError:
        temp_date = None
    return info


# =========================================================
# default column definitions
# =========================================================
col_a = Column(1, 1, 'Hull #', hull_space)
col_b = Column(2, 2, 'Boat Model', boat_model)
col_c = Column(3, 3, 'Order Details', order_details)
col_d = Column(4, 4, 'Colors    Interior / Exterior', colors_interior)
col_e = Column(5, 5, 'Engines', noop)
col_f = Column(6, 6, 'Current Phase', current_phase)
col_g = Column(7, 7, 'Est Start/Finish', start_finish)
col_g1 = Column(8, -1, 'Start Date', set_start_date)
col_g2 = Column(9, -1, 'Start Date', set_end_date)
col_h = Column(10, 8, 'Notes', noop)


# =========================================================
# clemens column defintions
# =========================================================
clm_a = Column(1, 1, 'Hull #', hull_space)
clm_b = Column(2, 2, 'Boat Model', boat_model)
clm_c = Column(3, 3, 'Package', noop)
clm_d = Column(4, 4, 'Colors    Interior / Exterior', colors_interior)
clm_e = Column(5, 5, 'Engines', noop)
clm_f = Column(6, 6, 'Order Details', order_details)
clm_g = Column(7, 7, 'Current Phase', current_phase)
clm_h = Column(8, 8, 'Est Start/Finish', start_finish)
col_h1 = Column(9, -1, 'Start Date', set_start_date)
col_h2 = Column(10, -1, 'Start Date', set_end_date)
clm_i = Column(11, 9, 'Notes', noop)


# =========================================================
# all dealer column defintions
# =========================================================
all_b = Column(3, 2, 'Boat Model', boat_model)
all_c = Column(4, 3, 'Package', noop)
all_d = Column(12, 4, 'Colors    Interior / Exterior', colors_interior)


# =========================================================
# dealership definitions
# =========================================================
reports = {
    'All Dealer': {
        'id': 2931819302676356,
        'name': 'All Dealer',
        'report': 'All Dealer - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            all_b,
            all_c,
            all_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Alaska Frontier Fabrication': {
        'id': 6215848202397572,
        'name': 'Alaska Frontier Fabrication',
        'report': 'Alaska Frontier Fabrication - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Avataa': {
        'id': 4374979853739908,
        'name': 'Avataa',
        'report': 'Avataa - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Boat Country': {
        'id': 1862555250517892,
        'name': 'Boat Country',
        'report': 'Boat Country - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Clemens Marina': {
        'id': 7685431392266116,
        'name': 'Clemens Marina',
        'report': 'Clemens Marina - Boats on Order',
        'template': 'BoatsOnOrderTemplateClemens.xlsx',
        'break1': 67,
        'break2': 72,
        'columns': [
            clm_a,
            clm_b,
            clm_c,
            clm_d,
            clm_e,
            clm_f,
            clm_g,
            col_h1,
            col_h2,
            clm_h,
            clm_i,
        ],
    },
    'Drummondville Marine': {
        'id': 8793087234336644,
        'name': 'Drummondville Marine',
        'report': 'Drummondville Marine - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Elephant Boys': {
        'id': 6603151173281668,
        'name': 'Elephant Boys',
        'report': 'Elephant Boys - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Enns Brothers': {
        'id': 8501389329491844,
        'name': 'Enns Brothers',
        'report': 'Enns Brothers - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Erie Marine Sales LLC': {
        'id': 6879464052287364,
        'name': 'Erie Marine Sales LLC',
        'report': 'Erie Marine Sales LLC - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Hampton Marine': {
        'id': 7533698787633028,
        'name': 'Hampton Marine',
        'report': 'Hampton Marine LLC. - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Port Boat House': {
        'id': 3591949602056068,
        'name': 'Port Boat House',
        'report': 'Port Boat House - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'The Bay Co': {
        'id': 4536017773455236,
        'name': 'The Bay Co',
        'report': 'The Bay Co - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Three Rivers': {
        'id': 7159452517328772,
        'name': 'Three Rivers',
        'report': 'Three Rivers - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Valley Marine': {
        'id': 875382787336068,
        'name': 'Valley Marine',
        'report': 'Valley Marine - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
    'Y Marina': {
        'id': 7940135837820804,
        'name': 'Y Marina',
        'report': 'Y Marina - Boats on Order',
        'template': 'BoatsOnOrderTemplate.xlsx',
        'break1': 64,
        'break2': 71,
        'columns': [
            col_a,
            col_b,
            col_c,
            col_d,
            col_e,
            col_f,
            col_g1,
            col_g2,
            col_g,
            col_h,
        ],
    },
}


# =========================================================
# helper functions
# =========================================================
def log(text, error=None):
    """
    print text to screen and make log to send by email in case of error
    """
    global log_text, errors
    print(text)
    log_text += text + "\n"
    if (error):
        errors = True


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def remove_files(*files):
    for file in files:
        with contextlib.suppress(FileNotFoundError):
            os.remove(file)

# =========================================================
# advanced date maniputlations
# =========================================================
def adjust_date(my_date):
    """
    roll date to first of the following month
    """
    if my_date.day >= rollover:
        while my_date.day > 1:
            my_date = my_date + datedelta.DAY
    return my_date


def start_info(value):
    """
    Convert dates like Jan, Jan 15, January 15, to January
    and Jan / Feb, Jan 15 / Feb 10, January 15 / February 10 to Jan / Feb
    and roll to next month if the date is past rollover in .env file

    If date gets past 3 months out then disable color change
    """
    global date_coloring
    output = ''
    text_color = '000000'  # default text color black
    dates = value.split('/')

    # process start date
    start = dateparser.parse(
        dates[0], settings={'PREFER_DATES_FROM': 'future'})

    # check for null start date
    if not start:
        return [output, text_color]

    # round up to the next month
    start_date = adjust_date(start)

    # disable date coloring if date is more than 4 months from now.
    if  start_date.month > [0,3,4,5,6,7,8,9,10,11,12,1,2,3,4,5,6][datetime.date.today().month]:
        date_coloring = False

    # Set colors for this month or next month
    if start_date.month == datetime.date.today().month and date_coloring:
        text_color = 'B00000'  # text color red
    elif start_date.month == (datetime.date.today() + datedelta.MONTH).month and date_coloring:
        text_color = '0000F0'  # text color blue
    # set output in case we are only outputting a start date
    output = start_date.strftime(one_date_fmt)

    # no end date
    if len(dates) == 1:
        return [output, text_color]

    # process end date
    end = dateparser.parse(dates[1], settings={'PREFER_DATES_FROM': 'future'})

    # check for null end date
    if not end:
        return [output, text_color]

    end_date = adjust_date(end)
    output = (start_date.strftime(two_date_fmt) +
              ' / ' + end_date.strftime(two_date_fmt))
    return [output, text_color]


# =========================================================
# headers / footers and cell boarder formatting
# =========================================================
def normal_border(dealer, row):
    """
    Normal row border with thicker far left and right lines
    """
    for i in range(1, len(dealer['columns']) - 1):
        side1 = 'thin'
        side2 = 'thin'
        if i == len(dealer['columns']) - 2:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        dealer['wsNew'].cell(column=i, row=row+dealer['base']).border = Border(
            right=Side(border_style=side1, color='FF000000'),
            left=Side(border_style=side2, color='FF000000'))


def side_border(dealer, row):
    """
    only far left and right sides get boarder
    """
    dealer['wsNew'].cell(column=1, row=row+dealer['base']).border = Border(
        left=Side(border_style='medium', color='FF000000'))
    dealer['wsNew'].cell(column=(len(dealer['columns'])-2),
                         row=row+dealer['base']).border = Border(
        right=Side(border_style='medium', color='FF000000'))


def heading_border(dealer, row):
    """
    write out header of column titles for all but the first page
    """
    for i in range(1, len(dealer['columns']) - 1):
        side1 = 'thin'
        side2 = 'thin'
        if i == len(dealer['columns']) - 2:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        dealer['wsNew'].cell(column=i, row=row+dealer['base']).border = Border(
            right=Side(border_style=side1, color='FF000000'),
            left=Side(border_style=side2, color='FF000000'),
            top=Side(border_style='medium', color='FF000000'),
            bottom=Side(border_style='medium', color='FF000000'))


def end_page_border(dealer, row):
    """
    bottom border after the last detail row
    """
    for i in range(1, len(dealer['columns']) - 1):
        side1 = 'thin'
        side2 = 'thin'
        if i == len(dealer['columns']) - 2:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        dealer['wsNew'].cell(column=i, row=row+dealer['base']).border = Border(
            right=Side(border_style=side1, color='FF000000'),
            left=Side(border_style=side2, color='FF000000'),
            bottom=Side(border_style='medium', color='FF000000'))


def bottom_border(dealer, row):
    """
    bottom border at the end of the footer
    """
    for i in range(1, len(dealer['columns']) - 1):
        side1 = 'thin'
        side2 = 'thin'
        if i == len(dealer['columns']) - 2:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        dealer['wsNew'].cell(column=i, row=row+dealer['base']).border = Border(
            right=Side(border_style=side1, color='FF000000'),
            left=Side(border_style=side2, color='FF000000'),
            bottom=Side(border_style='medium', color='FF000000'))


def set_mast_header(dealer, logo_name):
    """
    place logo, deaelername and date on first page of sheet
    """
    date = "Report Date: %s " % (
        datetime.datetime.today().strftime('%m/%d/%Y'))
    img = Image(logo_name)
    dealer['wsNew'].add_image(img, 'B1')
    dealer['wsNew']['B5'] = dealer['name']
    dealer['wsNew'].cell(column=(len(dealer['columns']) - 2), row=5).value = date


def set_header(dealer, row):
    """
    write column titles at top of page
    """
    heading_border(dealer, row)
    dealer['wsNew'].row_dimensions[row+dealer['base']].height = 21.6

    for column in dealer['columns']:
        if column.info['new'] == -1:
            continue
        dealer['wsNew'].cell(row=row+dealer['base'],
                             column=column.info['new'],
                             value=column.info['title'])
        dealer['wsNew'].cell(row=row+dealer['base'],
                             column=column.info['new']).alignment = Alignment(
                                 horizontal='center',
                                 vertical='center')
        dealer['wsNew'].cell(row=row+dealer['base'],
                             column=column.info['new']).font = Font(
                                 bold=True,
                                 size=9,
                                 name='Arial')


def set_footer(dealer, row):
    """
    footer for final page of report
    """
    side_border(dealer, row)
    side_border(dealer, row+1)

    dealer['wsNew'].merge_cells(start_row=row+dealer['base']+1,
                                start_column=1,
                                end_row=row+dealer['base']+1,
                                end_column=3)
    # insert left side footer here
    dealer['wsNew'].cell(row=row + dealer['base'] + 1,
                         column=1).alignment = Alignment(horizontal='center')
    dealer['wsNew'].cell(row=row + dealer['base'] + 1,
                         column=1).font = Font(bold=True)
    dealer['wsNew'].merge_cells(start_row=row+dealer['base']+2,
                                start_column=1,
                                end_row=row+dealer['base']+2,
                                end_column=(len(dealer['columns']) - 2))
    # insert center footer here 
    dealer['wsNew'].cell(row=row+dealer['base']+2,
                         column=1).alignment = Alignment(horizontal='center')
    dealer['wsNew'].cell(row=row+dealer['base']+2,
                         column=1).font = Font(bold=True)
    bottom_border(dealer, row+2)


# =========================================================
# add watermark to get colored footers on PDF
# =========================================================
def add_watermark(input, watermark, output):
    """
    add red and blue footer line to each page of the pdf
    will combine the watermark with the temp pdf and save
    as the final pdf
    """
    file = open(input, 'rb')
    reader = PdfFileReader(file)

    watermark = open(watermark, 'rb')
    reader2 = PdfFileReader(watermark)
    waterpage = reader2.getPage(0)

    writer = PdfFileWriter()

    for pageNum in range(0, reader.numPages):
        page = reader.getPage(pageNum)
        page.mergePage(waterpage)
        writer.addPage(page)

    resultFile = open(output, 'wb')
    writer.write(resultFile)
    file.close()
    resultFile.close()


# =========================================================
# process row and rows
# =========================================================
def fetch_value(cell):
    """
    fetch cell value and convert to data type that wont choke later functions
    """
    value = cell.value
    if cell.data_type == 's':
        return value
    if cell.is_date:
        return ('%02d/%02d/%02d' % (
            value.month,
            value.day,
            value.year - 2000))
    if value is None:
        return ''
    return str(int(value))


def process_row(dealer, row):
    """
    process one row by
      resetting default formatting
      read value of each column
      set font and color formatting rules for column
    after all columns have been collected
      render each cell font and background color
    """
    for column in dealer['columns']:
        column.reset()
        cell = dealer['wsOld'].cell(column=column.info['old'], row=row)
        column.info['text'] = fetch_value(cell)
        column.run()

    for column in dealer['columns']:
        if column.info['new'] == -1:
            continue
        cell = dealer['wsNew'].cell(column=column.info['new'],
                                    row=row+dealer['base']+dealer['offset'])
        cell.value = column.info['text']
        cell.font = column.font()
        cell.fill = PatternFill(start_color=column.bg(),
                                end_color=column.bg(),
                                fill_type="solid")


def process_rows(dealer, pdf):
    """
    Process all rows of sheet
    """
    dealer['pagelen'] = dealer['break1']
    dealer['page_number'] = 0
    dealer['offset'] = 0
    dealer['last_page_offset'] = 0
    i = 4

    for i in range(2, dealer['wsOld'].max_row + 1):
        process_row(dealer, i)
        # if there are not 3 lines left for footer on last page handle
        x = i > dealer['wsOld'].max_row - 4
        y = i > dealer['pagelen'] - dealer['base'] - 3
        if x and y and pdf:
            dealer['last_page_offset'] = 3
            dealer['pagelen'] = i + dealer['base']

        x = (i + dealer['base']) % dealer['pagelen'] == 0
        y = dealer['wsOld'].max_row != i
        if x and y and pdf:
            end_page_border(dealer, i + dealer['offset'])
            dealer['offset'] += 1 + dealer['last_page_offset']

            if i < dealer['wsOld'].max_row + 1:
                dealer['offset'] += 1
                set_header(dealer, i + dealer['offset'])

            dealer['page_number'] += 1
            dealer['pagelen'] += dealer['break2']
        else:
            normal_border(dealer, i + dealer['offset'])

    end_page_border(dealer, i + dealer['offset'])
    dealer['offset'] += 1
    set_footer(dealer, dealer['wsOld'].max_row + dealer['offset'])


# =========================================================
# process sheet to pdf or sheet to excel
# =========================================================
def process_sheet_to_pdf(dealer):
    """
    create pdf by
      creating temporary excel file
      load a librecalc template file to set page to landscape
      unoconvert excel file to temporary pdf file
      add watermark for red and blue text at bottom of page
      save resulting pdf with correct name in its final location
    """
    # change variables here
    input_file = resource_path(source_dir +
                               'downloads/' +
                               dealer['report'] +
                               '.xlsx')
    watermark_name = resource_path(source_dir +
                                   'watermark.pdf')
    temp_name = resource_path(source_dir + 'temp.xlsx')
    pdf_dir = (target_dir + 'Formatted - PDF/')
    output_name = pdf_dir + dealer['report'].replace('.','') + '.xlsx'
    logo_name = resource_path(source_dir + 'nrblogo1.jpg')
    dealer['base'] = 7

    # load sheet data is coming from
    wbOld = openpyxl.load_workbook(input_file)
    dealer['wsOld'] = wbOld.active

    # load sheet we are copying data to
    wbNew = openpyxl.load_workbook(resource_path(source_dir +
                                                 dealer['template']))
    dealer['wsNew'] = wbNew.active

    set_mast_header(dealer, logo_name)
    process_rows(dealer, True)
    range = 'A1:J'+str(dealer['wsNew'].max_row + 10)
    wbNew.create_named_range('_xlnm.Print_Area',
                             dealer['wsNew'], range, scope=0)

    # save new sheet out to temp.xls and temp.pdf file
    try:
        wbNew.save(output_name)
        result = subprocess.call(['/usr/bin/unoconv',
                                  '-f', 'pdf',
                                  '-t',
                                  resource_path(source_dir +
                                                'landscape.ots'),
                                  '--output=' + temp_name[:-4] + 'pdf',
                                  output_name])
        # change output_name to temp_name[:-4] + '1.xlsx'
        if (result):
            log('             UNICONV FAILED TO CREATE PDF', True)

    except Exception as e:
        log('             FAILED TO CREATE XLSX AND PDF: ' + str(e), True)

    # add watermark to temp.pdf and save to proper dealership name
    try:
        add_watermark(temp_name[:-4] + 'pdf',
                      watermark_name,
                      output_name[:-4] + 'pdf')
    except Exception as e:
        log('             FAILED TO ADD WATERMARK: ' + str(e), True)
    finally:
        remove_files(temp_name[:-4] + 'pdf')

def process_sheet_to_xlsx(dealer):
    """
    save excel file with correct name in its final location
    """
    # change variables here
    input_file = resource_path(source_dir +
                               'downloads/' +
                               dealer['report'] +
                               '.xlsx')
    output_name = target_dir + dealer['report'].replace('.','') + '.xlsx'
    logo_name = resource_path(source_dir + 'nrblogo1.jpg')
    dealer['base'] = 7

    # load sheet data is coming from
    wbOld = openpyxl.load_workbook(input_file)
    dealer['wsOld'] = wbOld.active

    # load sheet we are copying data to
    wbNew = openpyxl.load_workbook(resource_path(source_dir +
                                                 dealer['template']))
    dealer['wsNew'] = wbNew.active

    set_mast_header(dealer, logo_name)
    process_rows(dealer, False)
    range = 'A1:J'+str(dealer['wsNew'].max_row + 10)
    wbNew.create_named_range('_xlnm.Print_Area',
                             dealer['wsNew'],
                             range,
                             scope=0)

    # save new sheet out to new file
    try:
        wbNew.save(output_name)
    except Exception as e:
        log('             FAILED TO CREATE XLSX: ' + str(e), True)


def process_sheets(dealers, excel, pdf):
    """
    process all dealers by creating pdf and excel files as needed
    """
    global is_pdf
    global date_coloring 
    log("\nPROCESS SHEETS ===============================")
    if getattr(sys, 'frozen', False):
        os.chdir(resource_path(source_dir + 'downloads/'))
    for dlr in sorted(dealers):
        dealer = dealers[dlr]
        # check if file exists
        if pdf:
            is_pdf = True
            date_coloring = True
            log("  converting %s to pdf" % (dealer['report']))
            process_sheet_to_pdf(dealer)
            is_pdf = False
        if excel:
            log("  converting %s to xlsx" % (dealer['report']))
            process_sheet_to_xlsx(dealer)
        log("")


def download_sheets(dealers):
    """
    download excel spreadsheets via the smartsheet api for further processing
    """
    smart = smartsheet.Smartsheet(api)
    smart.assume_user(os.getenv('SMARTSHEET_USER'))
    log("DOWNLOADING SHEETS ===========================")
    for dlr in sorted(dealers):
        dealer = dealers[dlr]
        log("  downloading sheet: " + dealer['report'])
        try:
            smart.Reports.get_report_as_excel(dealer['id'],
                                              resource_path(source_dir)
                                              + 'downloads')
        except Exception as e:
            log('                     ERROR DOWNLOADING SHEET: ' +
                str(e), True)


def send_error_report():
    """
    used by try/except to send error report
    """
    subject = 'Smartsheet Boats on Order Error Report'
    mail_results(subject, "<pre>\n" + log_text + "</pre>\n")


def main(dealers, download, excel, pdf):
    """
    load environmental variables then download and process spreadsheets
    """
    global api, source_dir, target_dir, rollover, one_date_fmt, two_date_fmt
    global log_text, errors

    # load environmental variables
    env_path = resource_path('.env')
    load_dotenv(dotenv_path=env_path)

    log_text = ''
    errors = False
    api = os.getenv('SMARTSHEET_API')
    source_dir = os.getenv('SOURCE_DIR')
    target_dir = os.getenv('TARGET_DIR')
    rollover = int(os.getenv('ROLLOVER'))
    one_date_fmt = os.getenv('ONEDATEFMT')
    two_date_fmt = os.getenv('TWODATEFMT')

    # create download dir if necessary
    if not os.path.exists(resource_path(source_dir + 'downloads')):
        os.makedirs(resource_path(source_dir + 'downloads'))

    try:
        if download:
            download_sheets(dealers)
        if excel or pdf:
            process_sheets(dealers, excel, pdf)
    except Exception as e:
        log('Uncaught Error in main(): ' + str(e), True)

    if (errors):
        send_error_report()


@click.command()
@click.option(
    '--download/--no-download',
    default=True,
    help='Download sheet(s)'
)
@click.option(
    '--pdf/--no-pdf',
    default=True,
    help='Create PDFs'
)
@click.option(
    '--excel/--no-excel',
    default=True,
    help='Create Excel Sheets'
)
@click.option(
    '--dealer',
    '-d',
    multiple=True,
    help='Dealers to include'
)
@click.option(
    '--ignore',
    '-i',
    multiple=True,
    help='Dealers to ignore'
)
def cli(download, pdf, excel, dealer, ignore):
    """converts smartsheet boats on order report to
    excel sheets and pdf files for each dealership
    """

    dealers = {}
    # Add dealers we want to report on
    if dealer:
        for name in dealer:
            item = reports.get(name)
            if item:
                dealers[name] = item
    else:
        dealers = reports

    # Delete dealers we are not intested in
    if ignore:
        for name in ignore:
            if dealers.get(name):
                del dealers[name]

    main(dealers, download, excel, pdf)


if __name__ == "__main__":
    cli()  # pylint: disable=no-value-for-parameter
